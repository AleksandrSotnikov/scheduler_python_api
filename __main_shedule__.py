import json

import openpyxl


# Функция для парсинга Excel файла
def cell_value(sheet, row, column):
    value = sheet.cell(row=row, column=column).value
    return str(value) if value is not None else ""


def find_groups(sheet):
    groups = []
    for i in range(1, sheet.max_row):
        if cell_value(sheet, i, 2) == "1":
            group_name = cell_value(sheet, i, 1)
            groups.append((group_name, i))
    return groups


def check_pg(m1, m2, m3, m4):
    if m1 and not m2 and not m3 and m4:
        return "vmeste"
    elif m1 and m2 and not m3 and not m4:
        return "onlyone"
    elif not m1 and not m2 and m3 and m4:
        return "onlytwo"
    elif m1 and m2 and m3 and m4:
        return "onetwo"
    else:
        return "none"


def add_record(records, week, day, group, lesson_number, subgroup, subject, instructor, classroom):
    record = {
        "week_number": week,
        "day_of_week": day,
        "group_name": group,
        "lesson_number": lesson_number,
        "subgroup": subgroup,  # 0: Вместе, 1: Подгруппа 1, 2: Подгруппа 2
        "subject": subject,
        "instructor": instructor,
        "classroom": classroom
    }
    records.append(record)


def create_main_shedule(file_path, output_json):
    # Загружаем Excel файл
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    records = []
    weekbool = False
    week = 1
    day = 1

    # Получаем список всех групп с их строками начала
    groups = find_groups(sheet)

    for group_name, group_row in groups:
        for den in range(5, 64, 5):  # День
            for para in range(group_row, group_row + 12, 2):  # Пара (идёт через 2 строки)
                paranum = cell_value(sheet, para, 2)

                # Проверка: если значение нельзя преобразовать в целое число, продолжаем цикл
                try:
                    paranum = int(paranum)
                except ValueError:
                    continue  # Пропуск строки, если "Пара" или другое нечисловое значение

                pg = check_pg(
                    cell_value(sheet, para, den),
                    cell_value(sheet, para, den + 1),
                    cell_value(sheet, para, den + 2),
                    cell_value(sheet, para, den + 3)
                )

                if pg == "vmeste":
                    add_record(records, week, day, group_name, paranum, 0,
                               cell_value(sheet, para, den),
                               cell_value(sheet, para + 1, den),
                               cell_value(sheet, para, den + 3))
                elif pg == "onlyone":
                    add_record(records, week, day, group_name, paranum, 1,
                               cell_value(sheet, para, den),
                               cell_value(sheet, para + 1, den),
                               cell_value(sheet, para, den + 1))
                elif pg == "onlytwo":
                    add_record(records, week, day, group_name, paranum, 2,
                               cell_value(sheet, para, den + 2),
                               cell_value(sheet, para + 1, den + 2),
                               cell_value(sheet, para, den + 3))
                elif pg == "onetwo":
                    add_record(records, week, day, group_name, paranum, 1,
                               cell_value(sheet, para, den),
                               cell_value(sheet, para + 1, den),
                               cell_value(sheet, para, den + 1))
                    add_record(records, week, day, group_name, paranum, 2,
                               cell_value(sheet, para, den + 2),
                               cell_value(sheet, para + 1, den + 2),
                               cell_value(sheet, para, den + 3))

            if day == 6:
                weekbool = not weekbool
                week = 2 if weekbool else 1
            day = 1 if day == 6 else day + 1

    # Сохраняем в JSON файл
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump({"results": records}, f, ensure_ascii=False, indent=4)

    print(f"Результаты сохранены в {output_json}")