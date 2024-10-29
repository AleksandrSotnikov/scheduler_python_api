import json
import openpyxl
import sys
import os

def load_excel(file_path):
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' does not exist.")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        return sheet
    except Exception as e:
        print(f"Error loading Excel file: {str(e)}")
        sys.exit(1)

def cell_value(sheet, row, column):
    cell = sheet.cell(row=row, column=column)
    return cell.value if cell.value is not None else ""

def find_groups(sheet):
    groups = []
    for i in range(1, sheet.max_row):
        group_name = cell_value(sheet, i, 1)
        if group_name and group_name not in groups:
            groups.append((group_name, i))
    return groups

def check_pg(m1, m2, m3, m4):
    one = bool(m1)
    two = bool(m2)
    three = bool(m3)
    four = bool(m4)

    if not one and not two and not three and not four:
        return "none"
    elif one and not two and not three and four:
        return "vmeste"
    elif one and two and not three and not four:
        return "onlyone"
    elif not one and not two and three and four:
        return "onlytwo"
    elif one and two and three and four:
        return "onetwo"
    else:
        return "error"

def add_record(records, week, day, group, lesson_number, subgroup, subject, instructor, classroom):
    record = {
        "week_number": week,
        "day_of_week": day,
        "group_name": group,
        "lesson_number": lesson_number,
        "subgroup": subgroup,  # 0: Вместе, 1: Подгруппа 1, 2: Подгруппа 2
        "subject": subject,
        "instructor": instructor,
        "classroom": classroom
    }
    records.append(record)


def main(file_path, output_json):
    # Load Excel file
    sheet = load_excel(file_path)

    records = []
    weekbool = False
    week = 1
    day = 1

    # Получаем список всех групп с их строками начала
    groups = find_groups(sheet)

    for group_name, group_row in groups:
        for den in range(5, 64, 5):  # День
            for para in range(group_row, group_row + 12, 2):  # Пара (идёт через 2 строки)
                paranum = cell_value(sheet, para, 2)
                if not paranum:
                    continue  # Пропуск, если номер пары пуст

                try:
                    paranum = int(paranum)
                except ValueError:
                    continue

                pg = check_pg(
                    cell_value(sheet, para, den),
                    cell_value(sheet, para, den + 1),
                    cell_value(sheet, para, den + 2),
                    cell_value(sheet, para, den + 3)
                )

                if pg == "vmeste":
                    add_record(records, week, day, group_name, paranum, 0,
                               cell_value(sheet, para, den),
                               cell_value(sheet, para + 1, den),
                               cell_value(sheet, para, den + 3))
                elif pg == "onlyone":
                    add_record(records, week, day, group_name, paranum, 1,
                               cell_value(sheet, para, den),
                               cell_value(sheet, para + 1, den),
                               cell_value(sheet, para, den + 1))
                elif pg == "onlytwo":
                    add_record(records, week, day, group_name, paranum, 2,
                               cell_value(sheet, para, den + 2),
                               cell_value(sheet, para + 1, den + 2),
                               cell_value(sheet, para, den + 3))
                elif pg == "onetwo":
                    add_record(records, week, day, group_name, paranum, 1,
                               cell_value(sheet, para, den),
                               cell_value(sheet, para + 1, den),
                               cell_value(sheet, para, den + 1))
                    add_record(records, week, day, group_name, paranum, 2,
                               cell_value(sheet, para, den + 2),
                               cell_value(sheet, para + 1, den + 2),
                               cell_value(sheet, para, den + 3))

            if day == 6:
                weekbool = not weekbool
                week = 2 if weekbool else 1
            day = 1 if day == 6 else day + 1

    # Save to JSON file
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump({"results": records}, f, ensure_ascii=False, indent=4)

    print(f"Results have been saved to {output_json}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python shedule_parser.py <ExcelFilePath> <OutputJsonFile>")
        sys.exit(1)

    input_excel = sys.argv[1]
    output_json = sys.argv[2]

    main(input_excel, output_json)
